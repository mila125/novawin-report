from pywinauto import Application
import time
import os
import traceback
import pandas as pd
import configparser
import openpyxl
def manejar_novawin(path_novawin, archivo_qps):
    try:
        # Invertir las barras en la ruta del archivo
        archivo_qps = archivo_qps.replace("/", "\\")  # Reemplazar barras normales por barras invertidas
        
        # O usar normpath para normalizar la ruta según el sistema operativo
        archivo_qps = os.path.normpath(archivo_qps)

        # Inicializar NovaWin
        app, main_window = inicializar_novawin(path_novawin)

        # Interactuar con NovaWin
        seleccionar_menu(main_window, "File->Open")
        dialog = app.window(class_name="#32770")
        interactuar_con_cuadro_dialogo(dialog, archivo_qps)

        return app, main_window

    except Exception as e:
        print(f"Error al manejar NovaWin: {e}")
        traceback.print_exc()
        raise

def inicializar_novawin(path_novawin):
    try:
        app = Application(backend="uia").start(path_novawin)
        time.sleep(5)  # Esperar que se cargue NovaWin
        main_window = app.window(title_re=".*NovaWin.*")
        return app, main_window
    except Exception as e:
        print(f"Error al inicializar NovaWin: {e}")
        raise

def seleccionar_menu(window, ruta_menu):
    try:
        window.menu_select(ruta_menu)
        time.sleep(2)
    except Exception as e:
        print(f"Error al seleccionar menú '{ruta_menu}': {e}")
        raise

def interactuar_con_cuadro_dialogo(dialog, archivo):
    try:
        edit_box = dialog.child_window(class_name="Edit")
        edit_box.set_edit_text(archivo)
        open_button = dialog.child_window(class_name="Button", found_index=0)
        open_button.click_input()
    except Exception as e:
        print(f"Error al interactuar con el cuadro de diálogo: {e}")
        raise

def exportar_reporte(main_window, path_csv, app):
    try:
        # Buscar ventana de contexto y realizar exportación
        main_window.right_click_input()
        time.sleep(1)

        context_menu = app.window(title_re=".*Context.*")
        savecsv_menu_item = context_menu.child_window(title="Export to .CSV", control_type="MenuItem")
        savecsv_menu_item.click_input()

        time.sleep(2)
        csv_dialog = app.window(class_name="#32770")
        edit_box = csv_dialog.child_window(control_type="Edit", found_index=0)
        edit_box.set_edit_text(path_csv)
        save_button = csv_dialog.child_window(title="Guardar", control_type="Button")
        save_button.click_input()

        print("Archivo exportado exitosamente.")
    except Exception as e:
        print(f"Error durante la exportación: {e}")
        traceback.print_exc()
        raise

def leer_csv_y_crear_dataframe(ruta_csv):
    if not os.path.exists(ruta_csv):
        print(f"Archivo CSV no encontrado: {ruta_csv}")
        raise FileNotFoundError(f"Archivo no encontrado: {ruta_csv}")

    try:
        return pd.read_csv(ruta_csv)
    except Exception as e:
        print(f"Error al leer CSV: {e}")
        raise
def agregar_csv_a_plantilla_excel(ruta_csv, ruta_excel,df_csv):
    """
    Agrega el contenido de un CSV a una plantilla Excel (`Reporte.xlsx`).
    Los datos se escriben en las columnas vacías sin borrar el contenido existente.
    """
    try:
        # Invertir las barras en la ruta del archivo
        ruta_excel = ruta_excel.replace("/", "\\")  # Reemplazar barras normales por barras invertidas
        
        # O usar normpath para normalizar la ruta según el sistema operativo
        ruta_excel = os.path.normpath(ruta_excel)
        ruta_excel = os.path.join(ruta_excel, "Report.xlsx")
        print(ruta_excel)
        # Crear archivo Excel si no existe o si el archivo no tiene formato válido
        if not os.path.exists(ruta_excel) or not ruta_excel.endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"
            wb.save(ruta_excel)
            print(f"Archivo Excel creado: {ruta_excel}")

        # Abrir el archivo Excel
        wb = openpyxl.load_workbook(ruta_excel)
        if "Reporte" not in wb.sheetnames:
            ws = wb.create_sheet(title="Reporte")
        else:
            ws = wb["Reporte"]

        # Obtener la fila y columna inicial para insertar los datos
        max_row = ws.max_row
        max_col = ws.max_column

        # Determinar la columna vacía para comenzar a escribir
        start_col = max_col + 1 if max_row > 1 else 1

        # Escribir encabezados si es la primera inserción
        if start_col == 1:
            for col, header in enumerate(df_csv.columns, start=start_col):
                ws.cell(row=1, column=col).value = header

        # Insertar datos en columnas vacías
        for i, row in enumerate(df_csv.itertuples(index=False), start=2):
            for j, value in enumerate(row, start=start_col):
                ws.cell(row=i, column=j).value = value

        # Guardar cambios en el archivo Excel
        wb.save(ruta_excel)
        print(f"Datos del CSV agregados exitosamente a: {ruta_excel}")

    except Exception as e:
        print(f"Error al agregar datos del CSV a la plantilla Excel: {e}")
        raise
def guardar_dataframe_en_ini(df, archivo_ini):
    try:
        config = configparser.ConfigParser()
        for columna in df.columns:
            config[columna] = {f"fila_{i}": str(valor) for i, valor in enumerate(df[columna])}
        with open(archivo_ini, 'w') as archivo:
            config.write(archivo)
        print(f"DataFrame guardado en {archivo_ini}")
    except Exception as e:
        print(f"Error al guardar INI: {e}")
        raise