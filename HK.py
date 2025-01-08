import openpyxl
import os
import csv
import traceback
from datetime import datetime
from pywinauto import Application, findwindows
import time
import threading
from novawinmng import manejar_novawin, leer_csv_y_crear_dataframe,agregar_csv_a_plantilla_excel, guardar_dataframe_en_ini
from pywinauto.keyboard import send_keys

def generar_nombre_unico(base_path):
    # Normalizar las barras a formato Unix (/)
    base_path = base_path.replace("\\", "/")
    
    if not base_path.endswith("hk.csv"):
        base_path += "hk.csv"
    
    counter = 1
    while os.path.exists(base_path):
        name, ext = os.path.splitext(base_path)
        base_path = f"{name}_{counter}{ext}"
        counter += 1
    
    # Normalizar las barras de regreso a formato Windows (\)
    return base_path.replace("/", "\\")
def hilo_exportar(main_window, path_csv, app):
    # Aquí va la lógica para exportar el reporte
    exportar_reporte(main_window, path_csv, app)
def exportar_reporte(main_window, ruta_exportacion, app):
    try:
        # Imprimir detalles de los controles y ventanas hijas de main_window
        all_controls = main_window.children()
        main_window.print_control_identifiers()

        # Buscar el control por su clase
        graph_view_window = main_window.child_window(class_name="TGraphViewWindow")

        if graph_view_window.exists():
            print("Componente 'TGraphViewWindow' encontrado.")
            graph_view_window.print_control_identifiers()
            graph_view_window.right_click_input()
            time.sleep(1)
        else:
            print("No se encontró el componente 'TGraphViewWindow'.")

        context_menu = app.window(title_re=".*Context.*")
        tables_menu_item = context_menu.child_window(title="Tables", control_type="MenuItem")
        tables_menu_item.click_input()
        print("Menú 'Tables' seleccionado.")

        HK_method_menu_item = app.window(best_match="Tables").child_window(title="HK method", control_type="MenuItem")
        HK_method_menu_item.click_input()
        print("Submenú 'HK method' seleccionado.")

        HK_method_menu_item = app.window(best_match="HK method")
        HK_method_menu_item.print_control_identifiers(depth=2)
        Pore_Size_Distribution_menu_item = HK_method_menu_item.child_window(title=" Pore Size Distribution", control_type="MenuItem")
        Pore_Size_Distribution_menu_item.click_input()
        print("Se seleccionó ' Pore Size Distribution' exitosamente.")

        time.sleep(2)
        secondary_window2 = app.window(title_re=f".*tab:Pore Size Distribution: file_to_open_nameonly.*")
        main_window.right_click_input()
        time.sleep(1)

        savecsv_menu_item = context_menu.child_window(title="Export to .CSV", control_type="MenuItem")
        savecsv_menu_item.click_input()
        print("Se seleccionó 'Export to .CSV' exitosamente.")

        time.sleep(2)
        csv_dialog = app.window(class_name="#32770")

        print("llegó hasta aquí")
        ruta_exportacion = generar_nombre_unico(ruta_exportacion)

         # Enfocar el cuadro de texto con Alt + M
        send_keys('%m')  # % representa la tecla Alt en pywinauto
        time.sleep(2)
        send_keys(ruta_exportacion)  # % representa la tecla Alt en pywinauto
        # Esperar hasta que el cuadro de texto esté enfocado
        edit_box = csv_dialog.child_window(control_type="Edit", found_index=0)
       

        csv_button = csv_dialog.child_window(control_type="Button", title="Guardar") \
            if csv_dialog.child_window(control_type="Button", title="Guardar").exists() \
            else csv_dialog.child_window(control_type="Button", title="Save")
        
        if csv_button.exists():
            print("Existe el botón")
            csv_button.click_input()
            print("Archivo exportado exitosamente.")
            # Obtener ruta relativa
            ruta_relativa = os.path.relpath(ruta_exportacion, start=os.getcwd())
            print(f"Archivo exportado correctamente en: {ruta_relativa}")
            return ruta_relativa
        else:
            raise Exception("Botón 'Guardar' no encontrado.")

    except Exception as e:
        print(f"Error durante la exportación: {e}")
        traceback.print_exc()

def agregar_csv_a_excel(ruta_csv, ruta_excel):
    try:
        if not os.path.exists(ruta_excel):
            # Crear un nuevo archivo de Excel si no existe
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"
            wb.save(ruta_excel)
            print(f"Archivo Excel creado: {ruta_excel}")

        # Abrir el archivo de Excel existente
        wb = openpyxl.load_workbook(ruta_excel)
        if "Reporte" not in wb.sheetnames:
            ws = wb.create_sheet(title="Reporte")
        else:
            ws = wb["Reporte"]

        # Leer el archivo CSV
        with open(ruta_csv, "w", encoding="utf-8") as file:
            reader = csv.reader(file)
            for row in reader:
                ws.append(row)  # Agregar cada fila al Excel

        wb.save(ruta_excel)
        print(f"Reporte guardado exitosamente en: {ruta_excel}")

    except Exception as e:
        print(f"Error al agregar datos al archivo Excel: {e}")
        traceback.print_exc()
# Función para leer el CSV en un hilo
def hilo_leer_csv_y_crear_dataframe(ruta_csv, resultado_dict):
    try:
        resultado_dict['dataframe'] = leer_csv_y_crear_dataframe(ruta_csv)
    except Exception as e:
        resultado_dict['error'] = f"Error al leer CSV: {e}"

# Función para agregar el CSV al Excel en un hilo
def hilo_agregar_csv_a_plantilla_excel(ruta_csv, ruta_excel, resultado_dict):
    try:
        agregar_csv_a_plantilla_excel(ruta_csv, ruta_excel)
        resultado_dict['agregado'] = True
    except Exception as e:
        resultado_dict['error'] = f"Error al agregar datos del CSV a Excel: {e}"

# Función para guardar el DataFrame en un archivo INI en un hilo
def hilo_guardar_dataframe_en_ini(df, archivo_ini, resultado_dict):
    try:
        guardar_dataframe_en_ini(df, archivo_ini)
        resultado_dict['guardado'] = True
    except Exception as e:
        resultado_dict['error'] = f"Error al guardar INI: {e}"
def hk_main(path_qps, path_csv, path_novawin):
    print("Inicio de hk_main")
    try:
        # Inicializar y manejar NovaWin
        app, main_window = manejar_novawin(path_novawin, path_qps)

        # Exportar reporte y guardar la ruta del archivo exportado
        ruta_csv = exportar_reporte(main_window, path_csv, app)
        print(f"Archivo exportado a: {ruta_csv}")

        # Crear un hilo para la exportación (ya no es necesario exportar de nuevo)
        hilo_exportacion = threading.Thread(target=hilo_exportar, args=(main_window, path_csv, app))
        hilo_exportacion.start()

        # Esperar a que el hilo termine antes de proceder
        hilo_exportacion.join()

        # Mostrar el archivo exportado
        if ruta_csv:
            print(f"Archivo exportado exitosamente: {ruta_csv}")
        else:
            print("No se exportó ningún archivo.")

        # Crear DataFrame y guardar
        dataframe = leer_csv_y_crear_dataframe(ruta_csv)
        print(dataframe)
        agregar_csv_a_plantilla_excel(ruta_csv, path_csv,dataframe)
        guardar_dataframe_en_ini(dataframe, path_csv+"dataframe.ini")
        
        #dataframe.to_excel("reporte.xlsx", index=False, engine="openpyxl")

        print("Proceso completado exitosamente.")
        resultado_dict = {}
        # Crear hilos para cada tarea
        hilo_leer_csv = threading.Thread(target=hilo_leer_csv_y_crear_dataframe, args=(ruta_csv, resultado_dict))
        hilo_agregar_excel = threading.Thread(target=hilo_agregar_csv_a_plantilla_excel, args=(ruta_csv, path_csv, resultado_dict))
        hilo_guardar_ini = threading.Thread(target=hilo_guardar_dataframe_en_ini, args=(resultado_dict.get('dataframe', None), path_csv + "dataframe.ini", resultado_dict))

        # Iniciar hilos
        hilo_leer_csv.start()
        hilo_agregar_excel.start()
        hilo_guardar_ini.start()

        # Esperar a que todos los hilos terminen
        hilo_leer_csv.join()
        hilo_agregar_excel.join()
        hilo_guardar_ini.join()

        # Verificar errores o resultados en el diccionario
        if 'error' in resultado_dict:
            print(f"Error: {resultado_dict['error']}")
        else:
            print("Todas las tareas completadas exitosamente.")

        # Continuar con otras tareas si es necesario
        print("Proceso completado exitosamente.")

    except Exception as e:
        print(f"Error en hk_main: {e}")
        traceback.print_exc()

  